# -*- coding: UTF-8 -*-
"""
Author:Cheng Hong,Wu
去Line developer 中找到 Token 跟 user ID 執行後
透過CMD 連結到ngrok中
輸入 ngrok http 8888 找到下列連結網址
(執行失敗就嘗試重開電腦)
https://XXXXXX.XXXX.XXX.ngrok.io
並丟入LINE Develop 中的 webhook 並啟用

######功能設計######
1.功能介紹
2.幫忙找停車場
3.介紹停車場資訊(開放北區部分查詢)
"""
auth_token = ""
ownerID = ""

import requests
import random2 as random
import ParkingSearch as ps
import socketserver as socketserver
import http.server
from http.server import SimpleHTTPRequestHandler as RequestHandler
import json
from openpyxl import load_workbook

# 讀指令
wb = load_workbook('linebotcommad.xlsx')  # 讀取檔案
sheet = wb['command']
sheet_1 = wb['北區']

# 初始值
ReturnAnswer = [
            {
                "type": "text",
                "text": "抱歉無法判斷您想要問的事情，可撥打電話聯絡專人客服謝謝"
            }
        ]
def Message_Response(Input):
    commadAnswer="不好意思，我不清楚您想詢問什麼，可以點擊「選單」、點選中間的「如何使用」即有使用說明教學"
    locate = None
    # 讀基本指令
    for i in range(1, sheet.max_row + 1):
        if str(sheet.cell(row=i, column=2).value) == Input:
            commadAnswer = sheet.cell(row=i, column=3).value

    # 回傳停車場-區域
    if Input =="北區":
        commadAnswer = "北區目前開放的停車場有:\n"
        for i in range(1, sheet_1.max_column + 1):
            commadAnswer += str(sheet_1.cell(row=3, column=i).value) +"\n"
        commadAnswer +="請輸入該停車場名稱前面代碼(Ex:北區2)"

    for i in range(1, sheet_1.max_column + 1):
        if Input == str(sheet_1.cell(row=1, column=i).value):
            commadCall = str(sheet_1.cell(row=4, column=i).value)
            with open('park.json', encoding='utf-8') as data:
                data = json.load(data)
            for a in data:
                for b in a["ParkingLots"]:
                    if b["Position"] == commadCall:
                        commadAnswer = f"{b['Position']}\n總停車數:{b['TotalCar']}\n價格:\n{b['Notes']}"
                        locate = {
                          "type": "location",
                          "title": b['Position'],
                          "address":"點擊觀看位置並可連結Google導航",
                          "latitude": b["Y"],
                          "longitude": b["X"]
                        }
    ReturnAnswer = [
        {
            "type": "text",
            "text": commadAnswer
        }
    ]
    if not locate == None:
        ReturnAnswer.append(locate)
    return ReturnAnswer


class MyHandler(RequestHandler):
    def do_HEAD(self):
        self.send_response(200)
        self.end_headers()

    def do_POST(self):

        varLen = int(self.headers['Content-Length'])        # 取得讀取進來的網路資料長度
        if varLen > 0:
            post_data = self.rfile.read(varLen)             # 讀取傳過來的資料
            data = json.loads(post_data)                    # 把字串 轉成JSON
            replyToken = data['events'][0]['replyToken']      # 回傳要用Token

            # 使用者丟出地址
            if data['events'][0]['message']['type'] == "location":
                latitude = data['events'][0]['message']['latitude']         # 緯度
                longitude = data['events'][0]['message']['longitude']       # 經度

                # 開始找附近的停車場
                distance = "0m"
                parkData = ps.parkingLoading()
                nearestPark = ps.parkingSeach(parkData,longitude,latitude)
                # parkingName,parkingX,parkingY,parkingTotal,parkingAva,parkingNote,parkingUpadateTime,Distence
                price = nearestPark[5].replace('。',"\n")
                print(f"該地點的>>經度:{nearestPark[1]},緯度:{nearestPark[2]}")
                if nearestPark[7] < 1 :
                    distance = str(round((nearestPark[7]*1000),2))+"公尺"
                else:
                    distance = str(round(nearestPark[7],2)) + "公里"
                ReturnAnswer = [
                {
                  "type": "location",
                  "title": nearestPark[0],
                  "address":"點擊觀看位置並可連結Google導航",
                  "latitude": nearestPark[2],
                  "longitude": nearestPark[1]
                },{
                    "type": "text",
                    "text": f"{nearestPark[0]}\n總車位:{nearestPark[3]}個\n剩餘車位:{nearestPark[4]}個\n計價方式:\n{price}\n直線距離約為:\n{distance}\n\n更新時間:{nearestPark[6]}"

                  }
                ]

            # 文字訊息
            elif data['events'][0]['message']['type'] == "text":
                text = data['events'][0]['message']['text']       # 用戶的傳遞過來的文字內容
                print("text:", text)
                ReturnAnswer = Message_Response(text)

            # 回傳 發送者ID 跟訊息
            self.do_HEAD()
            message = {
                "replyToken": replyToken,
                "messages": ReturnAnswer
            }


            # 資料回傳 到 Line 的 https 伺服器

            hed = {'Authorization': 'Bearer ' + auth_token}
            url = 'https://api.line.me/v2/bot/message/reply'
            requests.post(url, json=message, headers=hed)         # 把資料HTTP POST送出去

socketserver.TCPServer.allow_reuse_address = True       # 可以重複使用IP
httpd = socketserver.TCPServer(('0.0.0.0', 8888), MyHandler)        # 啟動WebServer   :8888
httpd.serve_forever()